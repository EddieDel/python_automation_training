import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill


def extract_from_pdf(input):
    with pdfplumber.open(input) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()
        print(text)
        return text
    


def save_to_excel(text_out, prefered_name):
    wb = openpyxl.Workbook()
    ws = wb.active
   
    ws["A1"] = "MENU"
    lines = text_out.split("\n")
    rows = ["A"]
    Header_fonts = [ws["A1"]]


    print(f"Workbook Created")

    for x in rows:
        ws.column_dimensions[x].width = 90

    for i in Header_fonts:
        i.font = Font(bold=True)
        i.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for item in lines: 
        ws.append([item])

    wb.save(f"{prefered_name}.xlsx")
    print("Workbook Created")






sample_text = extract_from_pdf("sample-set-dinner-menus.pdf")
save_to_excel(sample_text, "extracted_info")    