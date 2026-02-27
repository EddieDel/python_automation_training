import requests
import openpyxl
from openpyxl.styles import Font, PatternFill

from bs4 import BeautifulSoup



def scrape_books(webpage):
    response = requests.get(webpage)
    response.encoding = "utf-8"
    books_list = []
    #make html readable
    soup = BeautifulSoup(response.text, "html.parser")
    soup.find_all("tag_name")


    #variable to store books
    books = soup.find_all("article", class_= "product_pod")


    #response status
    print(f"Scrapping was = {response.status_code}")
    #retrieve title
    print(f"The page title is: {soup.title}")

    #print number of books in that page
    y = len(books)
    print(f"We have: {y} books on the site")

    for item in books:
        z = item.find( "p", class_="price_color").text
        z = z.replace("£", "")
        z = float(z)
        x = item.find_all("a")[1]["title"]
        books_list.append({"title":x, "price":z})

    return books_list


def save_to_excel(list,prefered_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    total = 0

    ws["A1"] = "Title"
    ws["B1"] = "Price"



    columns = ["A", "B"]
    Header_fonts = [ws["A1"], ws["B1"]]


    print(f"Workbook Created")

    for x in columns:
        ws.column_dimensions[x].width = 20

    for i in Header_fonts:
        i.font = Font(bold=True)
        i.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for item in list: 
        ws.append([item["title"], item["price"]])
        total = total + item["price"]

    ws.append(["Total", total])    


    

    wb.save(f"{prefered_name}.xlsx")    
      
     


x = scrape_books("https://books.toscrape.com/")
save_to_excel(x,"scraped_book")
