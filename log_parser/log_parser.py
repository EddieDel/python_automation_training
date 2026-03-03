import openpyxl
from openpyxl.styles import Font, PatternFill


def read_log(filename):
    with open(filename, "r") as f:
        return f.readlines()
 

def parse_log(ex_lines):
    results = {"UVM_ERROR": [], "UVM_WARNING": [], "UVM_INFO": [], "UVM_FATAL": []}

    for item in ex_lines:
        if "UVM_INFO" in item:
            results["UVM_INFO"].append(item.strip())
        elif "UVM_ERROR" in item:
            results["UVM_ERROR"].append(item.strip())
        elif "UVM_WARNING" in item:
            results["UVM_WARNING"].append(item.strip())
        elif "UVM_FATAL" in item:
            results["UVM_FATAL"].append(item.strip())

    return results


def save_to_excel(results_out, prefered_name):
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["UVM_ERROR", "UVM_WARNING", "UVM_INFO", "UVM_FATAL"]

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        if header == "UVM_ERROR":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif header == "UVM_WARNING":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif header == "UVM_INFO":
            cell.fill = PatternFill(start_color="33FF00", end_color="33FF00", fill_type="solid")
        elif header == "UVM_FATAL":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 90

    # Write totals row
    for col, header in enumerate(headers, start=1):
        total = len(results_out[header])
        cell = ws.cell(row=2, column=col, value=f"TOTAL: {total}")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid")

    # Write log data starting from row 3
    max_len = max(len(results_out[h]) for h in headers)

    for row in range(max_len):
        for col, header in enumerate(headers, start=1):
            try:
                ws.cell(row=row + 3, column=col, value=results_out[header][row])
            except IndexError:
                pass

    wb.save(f"{prefered_name}.xlsx")
    print("Workbook Created")


lines = read_log("vcs_run.log")
parsed = parse_log(lines)
save_to_excel(parsed, "logoutput")

print(f"UVM_ERROR count: {len(parsed['UVM_ERROR'])}")
print(f"UVM_WARNING count: {len(parsed['UVM_WARNING'])}")
print(f"UVM_FATAL count: {len(parsed['UVM_FATAL'])}")
print(f"UVM_INFO count: {len(parsed['UVM_INFO'])}")