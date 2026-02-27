import openpyxl
from openpyxl.styles import Font, PatternFill


employees = [
    {"name":"John", "department":"IT", "salary": 3000},
    {"name":"Mike", "department":"Hardware", "salary": 4000},
    {"name":"Nick", "department":"software", "salary": 2000},
    {"name":"Jay", "department":"IT", "salary": 3500},
    {"name":"Jax", "department":"CEO", "salary": 5000}
]

employees2 = [
    {"name":"John2", "department":"IT", "salary": 3000},
    {"name":"Mike2", "department":"Hardware", "salary": 4000},
    {"name":"Nick2", "department":"software", "salary": 2000},
    {"name":"Jay2", "department":"IT", "salary": 3500},
    {"name":"Jax2", "department":"CEO", "salary": 5000}
]

def generate_report(employees_list,excel_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    total = 0

    ws["A1"] = "Name"
    ws["B1"] = "Department"
    ws["C1"] = "Salary"


    columns = ["A", "B", "C"]
    Header_fonts = [ws["A1"], ws["B1"], ws["C1"]]


    print(f"Workbook Created")

    for x in columns:
        ws.column_dimensions[x].width = 20

    for i in Header_fonts:
        i.font = Font(bold=True)
        i.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for item in employees_list: 
        ws.append([item["name"], item["department"], item["salary"]])
        total = total + item["salary"]

    ws.append(["Total", "", total])    


    

    wb.save(f"{excel_name}.xlsx")
    


generate_report(employees, "list1")
generate_report(employees2, "list2")